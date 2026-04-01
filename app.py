import streamlit as st
import os
import io
import json
import time
import pandas as pd
from PIL import Image
from dotenv import load_dotenv
from google import genai
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 1. 환경 설정 및 API 초기화
load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")

# 페이지 설정
st.set_page_config(
    page_title="영수증 자동 처리기",
    page_icon="🧾",
    layout="wide"
)

# 커스텀 CSS
st.markdown("""
<style>
    .reportview-container {
        background: #f0f2f6
    }
    .stButton>button {
        width: 100%;
        border-radius: 5px;
        height: 3em;
        background-color: #ff4b4b;
        color: white;
    }
    .stDataFrame {
        border-radius: 10px;
    }
</style>
""", unsafe_allow_html=True)

# 2. 세션 상태 초기화
if 'results' not in st.session_state:
    st.session_state.results = []
if 'failure_details' not in st.session_state:
    st.session_state.failure_details = []

# 3. 사이드바 구성
with st.sidebar:
    st.header("📖 사용 방법")
    st.markdown("""
    1. 왼쪽의 **파일 업로드** 섹션에서 영수증 사진을 선택하세요.
    2. 지원 파일: **JPG, JPEG, PNG, WEBP** (여러 장 동시 업로드 가능)
    3. **[분석 시작]** 버튼을 누르면 인식이 시작됩니다.
    4. 분석이 완료되면 결과 표와 함께 **엑셀 다운로드** 버튼이 나타납니다.
    """)
    st.divider()
    st.info("AI 모델: Gemini 3.1 Flash Lite (Preview)")
    
    # 세션 초기화 버튼 (새로운 분석을 위해)
    if st.button("🔄 분석 결과 초기화"):
        st.session_state.results = []
        st.session_state.failure_details = []
        st.rerun()

# 4. 메인 화면 구성
st.title("🧾 영수증 자동 처리기")
st.markdown("---")

if not api_key:
    st.error("경고: .env 파일에서 API 키를 찾을 수 없습니다. 설정을 확인해 주세요.")
    st.stop()

client = genai.Client(api_key=api_key)

# 파일 업로드 위젯
uploaded_files = st.file_uploader(
    "관리할 영수증 이미지를 업로드하세요 (다중 선택 가능)",
    type=["jpg", "jpeg", "png", "webp"],
    accept_multiple_files=True
)

if uploaded_files:
    st.write(f"📂 총 {len(uploaded_files)}개의 파일이 선택되었습니다.")
    
    if st.button("🚀 분석 시작"):
        # 분석 전 세션 데이터 초기화
        st.session_state.results = []
        st.session_state.failure_details = []
        
        # 진행 상태 표시
        progress_text = "작업 중입니다. 잠시만 기다려 주세요..."
        my_bar = st.progress(0, text=progress_text)
        
        prompt = """
이 영수증을 분석해서 아래 항목을 정확하게 추출해줘.

- 날짜: 영수증 발행일 (YYYY-MM-DD 형식)
- 상호명: 가게명 또는 건물명
- 공급가액: 부가세가 포함되지 않은 공급액 (영수증에 '공급가액' 또는 '공급대가' 또는 '소계' 또는 VAT제외 금액)
- 부가세: 부가가치세 금액 (0이면 0으로 표시)
- 카테고리: 상호명과 구매 항목을 보고 [식비, 교통비, 사무용품, 숙박비, 기타] 중 하나 선택

항목을 판독할 수 없는 경우만 '없음'으로 써.
반드시 아래 JSON 형식으로만 응답해 (숫자는 단위 없이 숫자만):
{
  "날짜": "YYYY-MM-DD 또는 없음",
  "상호명": "문자열 또는 없음",
  "공급가액": "숫자 또는 없음",
  "부가세": "숫자 (없으면 0)",
  "카테고리": "식비 또는 교통비 또는 사무용품 또는 숙박비 또는 기타"
}
"""
        
        for idx, uploaded_file in enumerate(uploaded_files):
            # 프로그레스 업데이트
            percent_complete = (idx + 1) / len(uploaded_files)
            my_bar.progress(percent_complete, text=f"Processing: {uploaded_file.name} ({idx+1}/{len(uploaded_files)})")
            
            try:
                # 이미지 로드
                img = Image.open(uploaded_file)
                
                # API 호출 (gemini-3.1-flash-lite-preview)
                response = client.models.generate_content(
                    model='gemini-3.1-flash-lite-preview',
                    contents=[prompt, img],
                    config={'response_mime_type': 'application/json'}
                )
                
                # 결과 파싱
                data = json.loads(response.text)
                
                # 실패 요건 체크 (기준: 날짜, 상호명, 공급가액 중 2개 이상 없음)
                check_fields = ["날짜", "상호명", "공급가액"]
                missing_count = sum(1 for f in check_fields if str(data.get(f, "없음")) == "없음")
                
                if missing_count >= 2:
                    data = {"파일명": uploaded_file.name, "날짜": "실패", "상호명": "실패", "공급가액": "실패", "부가세": "실패", "카테고리": "실패"}
                    st.session_state.failure_details.append({"파일명": uploaded_file.name, "사유": "이미지에서 정보 인식 부족"})
                else:
                    data['파일명'] = uploaded_file.name
                    # 카테고리 기본값 보정 (응답에 없을 경우)
                    if '카테고리' not in data:
                        data['카테고리'] = '기타'
                    
                st.session_state.results.append(data)
                
            except Exception as e:
                st.session_state.results.append({"파일명": uploaded_file.name, "날짜": "에러", "상호명": "에러", "공급가액": "에러", "부가세": "에러", "카테고리": "에러"})
                st.session_state.failure_details.append({"파일명": uploaded_file.name, "사유": str(e)})
            
            # API 할당량 고려 지연
            if len(uploaded_files) > 1:
                time.sleep(2)
        
        my_bar.empty()
        st.success("✅ 분석 완료!")
        st.rerun()  # 상태 반영을 위해 재실행

# 5. 결과 출력 섹션 (세션 데이터가 있는 경우 상시 노출)
if st.session_state.results:
    st.divider()
    df = pd.DataFrame(st.session_state.results)
    cols = ['파일명', '날짜', '상호명', '카테고리', '공급가액', '부가세', '부가세포함', '합계금액']
    # 누락 컬럼 기본값 대응
    for col in ['카테고리']:
        if col not in df.columns:
            df[col] = '없음'
    # 공급가액 컬럼이 없는 구 데이터 호환 처리
    if '공급가액' not in df.columns and '총금액' in df.columns:
        df['공급가액'] = df['총금액']

    # 부가세포함 판별: 부가세가 0보다 크면 "포함", 그 외 "미포함" ('실패', '에러'는 그대로 유지)
    def check_vat(x):
        if str(x) in ['실패', '에러']:
            return str(x)
        try:
            return '포함' if float(str(x).replace(',', '')) > 0 else '미포함'
        except Exception:
            return '미포함'
    df['부가세포함'] = df['부가세'].apply(check_vat)

    # 합계금액 계산: 공급가액 + 부가세
    def calc_total(row):
        val = str(row['공급가액'])
        if val in ['실패', '에러']:
            return val
        try:
            return int(float(val.replace(',', ''))) + int(float(str(row['부가세']).replace(',', '')))
        except Exception:
            return '-'
    df['합계금액'] = df.apply(calc_total, axis=1)
    df = df[cols]
    
    st.subheader("📊 분석 결과 목록")
    
    # 웹 화면 표시용 요약(합계) 행 추가
    total_val = pd.to_numeric(df['합계금액'], errors='coerce').sum()
    sum_row = {c: '' for c in cols}
    sum_row['파일명'] = '합계'
    sum_row['합계금액'] = int(total_val) if pd.notna(total_val) else 0
    df_display = pd.concat([df, pd.DataFrame([sum_row])], ignore_index=True)
    
    # 조건부 스타일링
    def highlight_failures(val):
        if val in ['실패', '에러']:
            return 'color: #ff4b4b; font-weight: bold;'
        return ''

    # 최신 Pandas(3.0 이상)에서는 applymap 대신 map을 사용합니다.
    try:
        styled_df = df_display.style.map(highlight_failures)
    except AttributeError:
        styled_df = df_display.style.applymap(highlight_failures)

    st.dataframe(styled_df, use_container_width=True)
    
    # 요약 정보
    fail_count = len(st.session_state.failure_details)
    success_count = len(st.session_state.results) - fail_count
    st.markdown(f"### ✅ 성공 {success_count}건 | ❌ 실패 {fail_count}건")
    
    # 엑셀 다운로드 (바이너리 생성)
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(cols)  # 헤더 행
    
    # ---- 숫자 정렬 맞춤 처리 ----
    # 문자로 된 숫자('103450' 등)를 진짜 숫자로 바꿔주어 표에서 오른쪽으로 예쁘게 정렬되도록 합니다.
    # 단, '실패', '없음' 등 문자는 무시하고 그대로 둡니다.
    for col in ['공급가액', '부가세']:
        # 쉼표를 제거하고 숫자로 강제 변환하되, 실패 시 문자열 원본을 그대로 유지합니다.
        clean_col = df[col].astype(str).str.replace(',', '')
        df[col] = pd.to_numeric(clean_col, errors='coerce').fillna(df[col])

    for _, row in df.iterrows():
        ws.append(list(row[cols]))
    
    # 총 합계액 계산 (화면에서 썼던 total_val을 그대로 사용)
    # openpyxl 수식(=SUM)은 인터넷 다운로드 파일(제한된 보기)에서 빈칸으로 보일 수 있어 실제 값을 삽입합니다.
    summary_row = [''] * len(cols)
    summary_row[0] = '합계'
    total_col_idx = cols.index('합계금액')
    summary_row[total_col_idx] = int(total_val) if pd.notna(total_val) else 0
    ws.append(summary_row)
    
    if st.session_state.failure_details:
        ws_fail = wb.create_sheet("실패 목록")
        ws_fail.append(["파일명", "실패 사유"])
        for f in st.session_state.failure_details:
            ws_fail.append([f["파일명"], f["사유"]])
            
    wb.save(output)
    
    st.download_button(
        label="📥 엑셀 파일 다운로드 (results.xlsx)",
        data=output.getvalue(),
        file_name="receipt_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    if not uploaded_files:
        st.info("분석할 영수증 파일을 업로드해 주세요.")
